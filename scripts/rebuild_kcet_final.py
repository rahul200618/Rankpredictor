#!/usr/bin/env python3
"""
KCET FINAL Comprehensive Data Extractor v2
============================================
Extracts ALL cutoff data from:
  - 2023 XLSX files (R1, R2, R3/Extended)  — MULTI-COLLEGE PER SHEET
  - 2024 XLSX files (Mock, R1, R2, R3/Extended) — MULTI-COLLEGE PER SHEET
  - 2025 PDF files (Mock, R1, R2, R3) — CROSS-PAGE TRACKING

Key fixes over v1:
  1. XLSX: Scans entire sheet for ALL college blocks (not just the first)
  2. PDF: Cross-page college header tracking
  3. Clean course names (no \\n characters)
  4. Proper dedup key (without cutoff_rank) — keep lowest rank
  5. All 24 categories captured
"""

import os
import re
import csv
import json
import sys
from pathlib import Path
from collections import defaultdict
from datetime import datetime

import openpyxl
import pdfplumber

# ─── Configuration ────────────────────────────────────────────────────────────

ROOT = Path(__file__).resolve().parent.parent
PUBLIC = ROOT / "public"
CUTOFFS_DIR = PUBLIC / "cutoffs"
DATA_DIR = PUBLIC / "data"

OUTPUT_MAIN = PUBLIC / "kcet_cutoffs.json"
OUTPUT_2023 = DATA_DIR / "cutoffs-2023.json"
OUTPUT_2024 = DATA_DIR / "cutoffs-2024.json"
OUTPUT_2025 = DATA_DIR / "cutoffs-2025.json"
OUTPUT_SUMMARY = DATA_DIR / "cutoffs-summary.json"
OUTPUT_CSV = DATA_DIR / "kcet_extracted_all.csv"

# All 24 KCET category codes
ALL_CATEGORIES = {
    '1G', '1K', '1R',
    '2AG', '2AK', '2AR',
    '2BG', '2BK', '2BR',
    '3AG', '3AK', '3AR',
    '3BG', '3BK', '3BR',
    'GM', 'GMK', 'GMR',
    'SCG', 'SCK', 'SCR',
    'STG', 'STK', 'STR',
}

# Load course code -> full name mapping
COURSE_MAPPING = {}
mapping_file = DATA_DIR / "course_mapping.json"
if mapping_file.exists():
    with open(mapping_file, "r", encoding="utf-8") as f:
        COURSE_MAPPING = json.load(f)
    print(f"  Loaded {len(COURSE_MAPPING)} course code mappings")


# ─── Utility Functions ────────────────────────────────────────────────────────

def clean(text):
    """Normalize whitespace, trim, remove control/NBSP chars."""
    if text is None:
        return ""
    s = str(text)
    s = s.replace('\xa0', ' ')  # NBSP
    s = re.sub(r'[\r\n\t]+', ' ', s)
    s = re.sub(r'\s+', ' ', s)
    return s.strip()


def clean_course_name(raw_name):
    """Clean course name: kill newlines, normalize whitespace."""
    if not raw_name:
        return ""
    s = str(raw_name)
    s = s.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ').replace('\xa0', ' ')
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def parse_round(filename):
    """Extract round from filename. R3 and EXT unified to R3."""
    lower = filename.lower()
    if 'mock' in lower:
        return 'MOCK'
    if 'round3' in lower or 'extended' in lower or 'ext' in lower:
        return 'R3'
    if 'round2' in lower:
        return 'R2'
    if 'round1' in lower:
        return 'R1'
    return 'R1'


def parse_year(filename):
    """Extract year from filename."""
    m = re.search(r'(202[0-9])', filename)
    return m.group(1) if m else None


def parse_cutoff(value):
    """Parse a cutoff rank value. Returns int or None."""
    if value is None:
        return None
    s = str(value).strip().replace(',', '').replace('\xa0', '')
    if s in ('', '--', '-', 'nan', 'None', 'NA', 'N/A', 'N.A.', '0'):
        return None
    s = re.sub(r'[^\d.]', '', s)
    if not s:
        return None
    try:
        n = int(float(s))
    except (ValueError, OverflowError):
        return None
    if n < 1 or n > 500000:
        return None
    return n


def make_key(row):
    """Dedup key — does NOT include cutoff_rank."""
    return (
        row.get('institute_code', ''),
        row.get('course', ''),
        row.get('category', ''),
        row.get('year', ''),
        row.get('round', ''),
    )


def resolve_course(raw):
    """Expand 2-3 letter codes to full names via mapping."""
    cleaned = clean_course_name(raw)
    if not cleaned:
        return cleaned
    parts = re.split(r'[\s\-]+', cleaned, maxsplit=1)
    code = parts[0].upper()
    if code in COURSE_MAPPING:
        return COURSE_MAPPING[code]
    if len(cleaned) <= 4 and cleaned.upper() in COURSE_MAPPING:
        return COURSE_MAPPING[cleaned.upper()]
    return cleaned


# ─── XLSX Extraction (2023 & 2024) ────────────────────────────────────────────
#
# CRITICAL: Each sheet has MULTIPLE colleges stacked vertically:
#   Row N:   E001  College Name
#   Row N+1: 1G  1K  1R  2AG ... STK STR   (category header)
#   Row N+2: AI Artificial  9507  18087 ...  (data row)
#   Row N+3: CE Civil       9920  ...        (data row)
#   ... more data rows ...
#   Row M:   E002  Next College Name         (next college block)
#   Row M+1: 1G  1K  1R  2AG ...             (category header)
#   ... and so on

def find_xlsx_files(years):
    """Find XLSX files, deduplicate by filename."""
    files = []
    seen_names = set()
    for search_dir in [PUBLIC, CUTOFFS_DIR, ROOT]:
        if not search_dir.exists():
            continue
        for f in sorted(search_dir.glob("kcet-*-cutoffs.xlsx")):
            fname = f.name.lower()
            if fname in seen_names:
                continue
            for y in years:
                if str(y) in fname:
                    files.append(f)
                    seen_names.add(fname)
                    break
    return files


def extract_xlsx_file(filepath, year, round_val):
    """
    Extract cutoff data from a SINGLE XLSX file.
    Handles MULTIPLE colleges per sheet.
    """
    results = []
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row is None or max_col is None:
            continue

        # Phase 1: Scan the entire sheet to find all college header rows
        # and all category header rows.
        college_rows = []   # [(row_idx, code, name), ...]
        category_rows = []  # [(row_idx, {cat: col_idx}), ...]

        for row_idx in range(1, max_row + 1):
            row_vals = {}
            row_text_parts = []
            cat_map = {}

            for col_idx in range(1, min(max_col + 1, 30)):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is None:
                    continue
                s = clean(v)
                if not s:
                    continue
                row_vals[col_idx] = s
                row_text_parts.append(s)

                # Check if this cell is a category header
                if s.upper() in ALL_CATEGORIES:
                    cat_map[s.upper()] = col_idx

            row_text = " ".join(row_text_parts)

            # Try to detect a college code in this row
            m = re.search(r'\b(E\d{3,4}[A-Z]?)\b', row_text)
            if m:
                code = m.group(1)
                # Extract the college name (text after the code)
                # Remove the header/title parts
                name_text = row_text
                # Remove the "ENGINEERING CUTOFF RANK..." prefix if present
                name_text = re.sub(r'ENGINEERING.*?ALLOTMENT.*?\)', '', name_text, flags=re.IGNORECASE)
                name_text = re.sub(r'\d{2}-[A-Z]{3}-\d{2}.*?(AM|PM)', '', name_text, flags=re.IGNORECASE)
                # Extract text after the code
                name_match = re.search(re.escape(code) + r'\s+(.*)', name_text)
                name = name_match.group(1).strip() if name_match else f"College {code}"
                # Clean leftover noise
                name = re.sub(r'^\d+\s*', '', name)  # Leading numbers
                name = name.strip()
                if not name:
                    name = f"College {code}"

                # Only consider this a college row if it has few data cells
                # (college header rows don't have 24+ numeric values)
                numeric_count = sum(1 for v in row_vals.values() if re.match(r'^\d+\.?\d*$', v))
                if numeric_count < 10:
                    college_rows.append((row_idx, code, name))

            # Check for category header row (must have many categories)
            if len(cat_map) >= 10:
                category_rows.append((row_idx, cat_map))

        # Phase 2: Build college blocks
        # Each college block = college row + closest category row + data rows until next college
        if not college_rows:
            continue

        for i, (college_row, code, name) in enumerate(college_rows):
            # Find the category header row closest to (and after) this college row
            cat_row = None
            cat_map = None
            for cr_idx, cr_map in category_rows:
                if cr_idx > college_row:
                    cat_row = cr_idx
                    cat_map = cr_map
                    break
            if cat_row is None or cat_map is None:
                continue

            # Data rows: from cat_row+1 until the next college row (or end of sheet)
            next_college_row = college_rows[i + 1][0] if i + 1 < len(college_rows) else max_row + 1

            # Find the course name column (first column with text in data rows)
            course_col = 1
            for col_idx in range(1, min(6, max_col + 1)):
                text_count = 0
                for row_idx in range(cat_row + 1, min(cat_row + 10, next_college_row)):
                    v = clean(ws.cell(row=row_idx, column=col_idx).value)
                    if v and len(v) > 1 and re.search(r'[A-Za-z]', v):
                        text_count += 1
                if text_count >= 2:
                    course_col = col_idx
                    break

            for row_idx in range(cat_row + 1, next_college_row):
                raw_course = clean(ws.cell(row=row_idx, column=course_col).value)
                if not raw_course:
                    continue
                if raw_course in ('--', '-', 'nan'):
                    continue
                low = raw_course.lower()
                if 'course' in low or 'branch' in low or 'programme' in low:
                    continue
                if re.match(r'^[\d\s\-\.]+$', raw_course):
                    continue

                course_name = resolve_course(raw_course)
                if not course_name:
                    continue

                for cat, cat_col in cat_map.items():
                    rank = parse_cutoff(ws.cell(row=row_idx, column=cat_col).value)
                    if rank is None:
                        continue

                    results.append({
                        'institute': name,
                        'institute_code': code,
                        'course': course_name,
                        'category': cat,
                        'cutoff_rank': rank,
                        'year': year,
                        'round': round_val,
                    })

    wb.close()
    return results


def extract_all_xlsx(years):
    """Extract from all XLSX files for the given years."""
    files = find_xlsx_files(years)
    all_results = []
    for filepath in files:
        fname = filepath.name
        year = parse_year(fname)
        round_val = parse_round(fname)
        if year not in years:
            continue
        print(f"  📂 {fname} → Year={year}, Round={round_val}")
        rows = extract_xlsx_file(filepath, year, round_val)
        # Count unique colleges
        colleges = set(r['institute_code'] for r in rows)
        print(f"     ✅ {len(rows):,} entries from {len(colleges)} colleges")
        all_results.extend(rows)
    return all_results


# ─── PDF Extraction (2025) ─────────────────────────────────────────────────────

def extract_pdf_file(pdf_path, year, round_val):
    """
    Extract cutoff data from a KCET PDF.
    
    KEY FIX: Tracks current_college across pages and uses pure coordinate-based
    extraction (X, Y) instead of find_tables() to prevent cell shifting.
    """
    results = []

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        current_code = None
        current_name = None
        last_cat_coords = {}

        for page_idx, page in enumerate(pdf.pages):
            if page_idx % 25 == 0:
                print(f"     Page {page_idx + 1}/{total_pages}...")

            words = page.extract_words(keep_blank_chars=False)
            if not words:
                continue
                
            # Group words by Y coordinate (with 2.0 pt tolerance)
            rows = {}
            for w in words:
                y = round(w['top'], 1)
                matched_y = None
                for existing_y in rows:
                    if abs(existing_y - y) <= 2.0:
                        matched_y = existing_y
                        break
                if matched_y is None:
                    matched_y = y
                    rows[matched_y] = []
                rows[matched_y].append(w)
                
            sorted_ys = sorted(rows.keys())
            
            for y in sorted_ys:
                row_words = sorted(rows[y], key=lambda x: x['x0'])
                text_full = " ".join([w['text'] for w in row_words])
                text_clean = clean(text_full)

                # 1. College Header
                m = re.search(
                    r'College\s*:?\s*[\(\[]?([A-Z0-9]{3,6})[\)\]]?\s*(.*)',
                    text_clean,
                    re.IGNORECASE
                )
                if m:
                    c = m.group(1).strip()
                    n = m.group(2).strip()
                    n = re.sub(r'^[\(\)]+|[\(\)]+$', '', n).strip()
                    if len(c) >= 3 and re.match(r'^E\d', c):
                        current_code = c
                        current_name = n
                        last_cat_coords = {}
                        continue

                if not current_code:
                    continue

                # 2. Category Header Row
                upper_words = [w['text'].upper().strip() for w in row_words]
                cat_count = sum(1 for w in upper_words if w in ALL_CATEGORIES)
                
                if cat_count >= 5 or (('COURSE' in text_clean.upper() or 'BRANCH' in text_clean.upper()) and cat_count >= 3):
                    cat_coords = {}
                    for w in row_words:
                        text = w['text'].upper().strip()
                        if text in ALL_CATEGORIES:
                            center_x = (w['x0'] + w['x1']) / 2.0
                            cat_coords[text] = center_x
                    if cat_coords:
                        last_cat_coords = cat_coords
                    continue

                # 3. Data Row
                if not last_cat_coords:
                    continue
                    
                min_cat_x = min(last_cat_coords.values())
                
                course_words = [w for w in row_words if w['x1'] < min_cat_x - 10]
                data_words = [w for w in row_words if w['x1'] >= min_cat_x - 10]

                if not course_words:
                    continue

                course_raw = " ".join([w['text'] for w in course_words])
                course_raw = clean_course_name(course_raw)

                if not course_raw:
                    continue
                if course_raw.upper() in ('COURSE', 'BRANCH', 'PROGRAMME', 'COURSE NAME', '--', '-', ''):
                    continue
                if re.match(r'^[\d\s\-\.]+$', course_raw):
                    continue
                if 'Course' in course_raw and 'GM' not in course_raw:
                    continue

                for w in data_words:
                    val = w['text'].strip()
                    rank = parse_cutoff(val)
                    if rank is None:
                        continue
                        
                    center_x = (w['x0'] + w['x1']) / 2.0
                    closest_cat = None
                    min_dist = float('inf')
                    
                    for cat, cat_x in last_cat_coords.items():
                        dist = abs(cat_x - center_x)
                        if dist < min_dist:
                            min_dist = dist
                            closest_cat = cat
                            
                    # Tolerate up to ~35 points of horizontal drift
                    if closest_cat and min_dist < 40:
                        results.append({
                            'institute': current_name,
                            'institute_code': current_code,
                            'course': course_raw,
                            'category': closest_cat,
                            'cutoff_rank': rank,
                            'year': year,
                            'round': round_val,
                        })

    return results


def extract_pdfs_for_year(year):
    """Extract from all PDF files for a given year."""
    pdf_files = sorted(CUTOFFS_DIR.glob(f"kcet-{year}*.pdf"))
    all_results = []
    for pdf_path in pdf_files:
        fname = pdf_path.name
        round_val = parse_round(fname)
        print(f"  📄 {fname} → Year={year}, Round={round_val}")
        rows = extract_pdf_file(pdf_path, year, round_val)
        colleges = set(r['institute_code'] for r in rows)
        print(f"     ✅ {len(rows):,} entries from {len(colleges)} colleges")
        all_results.extend(rows)
    return all_results


# ─── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 70)
    print("  KCET FINAL COMPREHENSIVE DATA EXTRACTOR v2")
    print("  2023/2024 from XLSX (multi-college) | 2025 from PDFs")
    print("=" * 70)
    print()

    all_raw = []

    # ── Phase 1: 2023 XLSX ──
    print("📥 Phase 1: Extracting 2023 XLSX data...")
    xlsx_2023 = extract_all_xlsx(years=('2023',))
    colleges_2023 = set(r['institute_code'] for r in xlsx_2023)
    print(f"   → 2023 raw: {len(xlsx_2023):,} entries from {len(colleges_2023)} colleges")
    all_raw.extend(xlsx_2023)
    print()

    # ── Phase 2: 2024 XLSX ──
    print("📥 Phase 2: Extracting 2024 XLSX data...")
    xlsx_2024 = extract_all_xlsx(years=('2024',))
    colleges_2024 = set(r['institute_code'] for r in xlsx_2024)
    print(f"   → 2024 raw: {len(xlsx_2024):,} entries from {len(colleges_2024)} colleges")
    all_raw.extend(xlsx_2024)
    print()

    # ── Phase 3: 2025 PDF ──
    print("📥 Phase 3: Extracting 2025 PDF data...")
    pdf_2025 = extract_pdfs_for_year('2025')
    colleges_2025 = set(r['institute_code'] for r in pdf_2025)
    print(f"   → 2025 raw: {len(pdf_2025):,} entries from {len(colleges_2025)} colleges")
    all_raw.extend(pdf_2025)
    print()

    # ── Deduplication ──
    print("🔄 Deduplicating (key = institute_code+course+category+year+round)...")
    seen = {}
    dupes = 0
    for row in all_raw:
        key = make_key(row)
        if key in seen:
            dupes += 1
            if row['cutoff_rank'] < seen[key]['cutoff_rank']:
                seen[key] = row
        else:
            seen[key] = row

    all_data = list(seen.values())
    print(f"   Total raw: {len(all_raw):,}")
    print(f"   Duplicates removed: {dupes:,}")
    print(f"   Final unique: {len(all_data):,}")
    print()

    # ── Statistics ──
    by_year = defaultdict(int)
    by_round = defaultdict(int)
    by_year_round = defaultdict(int)
    inst_by_year = defaultdict(set)
    courses = set()
    cats = set()
    cat_counts = defaultdict(int)

    for r in all_data:
        by_year[r['year']] += 1
        by_round[r['round']] += 1
        by_year_round[f"{r['year']}-{r['round']}"] += 1
        inst_by_year[r['year']].add(r['institute_code'])
        courses.add(r['course'])
        cats.add(r['category'])
        cat_counts[r['category']] += 1

    # ── Validation ──
    print("🔍 Validation:")
    pres = [r for r in all_data if r['institute_code'] == 'E237']
    print(f"   Presidency (E237): {len(pres)} entries")
    for k in sorted(set(f"{r['year']}-{r['round']}" for r in pres)):
        cnt = sum(1 for r in pres if f"{r['year']}-{r['round']}" == k)
        print(f"     {k}: {cnt}")

    bad = [r for r in all_data if '\n' in r.get('course', '')]
    print(f"   Courses with \\n: {len(bad)} {'✅' if not bad else '❌'}")

    for yr in sorted(by_year):
        yr_cats = set(r['category'] for r in all_data if r['year'] == yr)
        print(f"   {yr} categories: {len(yr_cats)} {'✅' if len(yr_cats) >= 24 else '⚠️'}")

    for yr_rd in sorted(by_year_round):
        yr, rd = yr_rd.split('-', 1)
        colleges = set(r['institute_code'] for r in all_data if r['year'] == yr and r['round'] == rd)
        print(f"   {yr_rd}: {by_year_round[yr_rd]:>6,} entries, {len(colleges)} colleges")
    print()

    # ── Sort ──
    all_data.sort(key=lambda r: (r['year'], r['round'], r['institute_code'], r['course'], r['category']))

    # ── Save ──
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    data_2023 = [r for r in all_data if r['year'] == '2023']
    data_2024 = [r for r in all_data if r['year'] == '2024']
    data_2025 = [r for r in all_data if r['year'] == '2025']

    for fp, yd, label in [(OUTPUT_2023, data_2023, '2023'), (OUTPUT_2024, data_2024, '2024'), (OUTPUT_2025, data_2025, '2025')]:
        with open(fp, 'w', encoding='utf-8') as f:
            json.dump(yd, f, ensure_ascii=False)
        mb = fp.stat().st_size / (1024*1024)
        print(f"💾 {fp.name}: {len(yd):,} entries ({mb:.1f} MB)")

    metadata = {
        'last_updated': datetime.now().isoformat(),
        'source_type': 'comprehensive_extraction_v2',
        'total_entries': len(all_data),
        'total_institutes': len(set(r['institute_code'] for r in all_data)),
        'total_courses': len(courses),
        'total_categories': len(cats),
        'years_covered': sorted(by_year.keys(), reverse=True),
        'rounds_covered': sorted(by_round.keys()),
        'records_by_year': dict(sorted(by_year.items())),
        'records_by_round': dict(sorted(by_round.items())),
        'dedupe_key': 'institute_code+course+category+year+round',
    }
    output = {'metadata': metadata, 'cutoffs': all_data}
    with open(OUTPUT_MAIN, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False)
    mb = OUTPUT_MAIN.stat().st_size / (1024*1024)
    print(f"💾 {OUTPUT_MAIN.name}: {len(all_data):,} entries ({mb:.1f} MB)")

    # Copy to fallback paths
    for name in ['kcet_cutoffs_consolidated.json', 'kcet_cutoffs_master.json', 'kcet_cutoffs_high_volume.json']:
        dst = DATA_DIR / name
        with open(dst, 'w', encoding='utf-8') as f:
            json.dump(output, f, ensure_ascii=False)
    print(f"💾 Copied to 3 fallback paths in data/")

    summary = {
        'totals': {'records': len(all_data), 'colleges': len(set(r['institute_code'] for r in all_data)), 'branches': len(courses)},
        'years': dict(sorted(by_year.items())),
        'categories': dict(sorted(cat_counts.items(), key=lambda x: -x[1])),
    }
    with open(OUTPUT_SUMMARY, 'w', encoding='utf-8') as f:
        json.dump(summary, f, indent=2, ensure_ascii=False)
    print(f"💾 {OUTPUT_SUMMARY.name}")

    fieldnames = ['institute', 'institute_code', 'course', 'category', 'cutoff_rank', 'year', 'round']
    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_data)
    mb = OUTPUT_CSV.stat().st_size / (1024*1024)
    print(f"💾 {OUTPUT_CSV.name}: ({mb:.1f} MB)")
    print()

    # ── Final Summary ──
    print("=" * 70)
    print("  EXTRACTION SUMMARY v2")
    print("=" * 70)
    print(f"  📝 Total Entries:     {len(all_data):,}")
    print(f"  🏫 Unique Institutes: {len(set(r['institute_code'] for r in all_data))}")
    print(f"  📚 Unique Courses:    {len(courses)}")
    print(f"  📊 Categories:        {len(cats)}")
    print()
    for y, c in sorted(by_year.items()):
        ncol = len(inst_by_year[y])
        print(f"  {y}: {c:>8,} entries | {ncol} colleges")
    print()
    for r, c in sorted(by_round.items()):
        print(f"  {r:>4}: {c:>8,}")
    print("=" * 70)


if __name__ == "__main__":
    main()
