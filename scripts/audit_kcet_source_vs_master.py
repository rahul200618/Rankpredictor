#!/usr/bin/env python3
"""
Deep audit: compare extracted KCET master JSON against raw source files.

Sources:
- 2023/2024 XLSX files in public/
- 2025 PDFs in public/cutoffs/

Output:
- reports/kcet_audit_report.json
- reports/kcet_audit_missing.csv
- reports/kcet_audit_extra.csv
"""

from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import openpyxl
import pdfplumber


ROOT = Path(__file__).resolve().parent.parent
PUBLIC_DIR = ROOT / "public"
REPORTS_DIR = ROOT / "reports"

MASTER_JSON = PUBLIC_DIR / "data" / "kcet_cutoffs_master.json"
XLSX_FILES = sorted(PUBLIC_DIR.glob("kcet-202[34]*-cutoffs.xlsx"))
PDF_2025_FILES = sorted(PUBLIC_DIR.glob("cutoffs/kcet-2025*.pdf"))

CATEGORY_CODES = {
    "1G",
    "1K",
    "1R",
    "2AG",
    "2AK",
    "2AR",
    "2BG",
    "2BK",
    "2BR",
    "3AG",
    "3AK",
    "3AR",
    "3BG",
    "3BK",
    "3BR",
    "GM",
    "GMK",
    "GMR",
    "GMP",
    "NRI",
    "OPN",
    "OTH",
    "SCG",
    "SCK",
    "SCR",
    "STG",
    "STK",
    "STR",
}

_RE_SPACES = re.compile(r"\s+")
_RE_ECODE = re.compile(r"\b(E\d{3})\b", re.IGNORECASE)
_RE_NUMERIC = re.compile(r"^[\d.,\-\s]+$")


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ")
    text = text.replace("\n", " ").replace("\r", " ").replace("\t", " ")
    text = _RE_SPACES.sub(" ", text).strip()
    return text


def parse_rank(value: object) -> Optional[int]:
    text = clean_text(value)
    if text in {"", "-", "--", "NA", "N/A", "None", "nan"}:
        return None
    text = text.replace(",", "")
    text = re.sub(r"\s*\.\s*", ".", text)
    text = re.sub(r"(?<=\d)\s+(?=\d)", "", text)
    match = re.search(r"\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        rank = int(float(match.group(0)))
    except ValueError:
        return None
    if 1 <= rank <= 500000:
        return rank
    return None


def round_from_xlsx_filename(name: str) -> str:
    lower = name.lower()
    if "mock" in lower:
        return "MOCK"
    if "round3" in lower or "extended" in lower:
        return "R3"
    if "round2" in lower:
        return "R2"
    return "R1"


def year_from_filename(name: str) -> Optional[str]:
    match = re.search(r"(2023|2024|2025)", name)
    return match.group(1) if match else None


def round_from_pdf_filename(name: str) -> str:
    lower = name.lower()
    if "mock" in lower:
        return "MOCK"
    if "round3" in lower:
        return "R3"
    if "round2" in lower:
        return "R2"
    return "R1"


def looks_like_non_course(text: str) -> bool:
    upper = text.upper()
    if upper in CATEGORY_CODES:
        return True
    if _RE_ECODE.search(upper):
        return True
    if _RE_NUMERIC.match(text):
        return True
    if text in {"-", "--"}:
        return True
    blocked = (
        "ENGINEERING CUTOFF",
        "ALLOTMENT",
        "COURSE NAME",
        "COURSE",
        "BRANCH NAME",
        "BRANCH",
    )
    return any(token in upper for token in blocked)


def extract_category_map(row_values: List[object]) -> Dict[int, str]:
    category_map: Dict[int, str] = {}
    for idx, value in enumerate(row_values):
        category = clean_text(value).upper()
        if category in CATEGORY_CODES:
            category_map[idx] = category
    return category_map if len(category_map) >= 5 else {}


def extract_course_name(row_values: List[object], category_cols: Iterable[int]) -> Optional[str]:
    cols = sorted(category_cols)
    if not cols:
        return None
    first_cat = cols[0]

    for idx in range(0, min(first_cat, 10)):
        text = clean_text(row_values[idx] if idx < len(row_values) else "")
        if not text or looks_like_non_course(text):
            continue
        if re.search(r"[A-Za-z]", text):
            return text
    return None


def extract_college_from_row(
    row_values: List[object], known_names: Dict[str, str]
) -> Optional[Tuple[str, str]]:
    for idx, value in enumerate(row_values):
        text = clean_text(value)
        if not text:
            continue
        match = _RE_ECODE.search(text)
        if not match:
            continue
        code = match.group(1).upper()
        trailing = clean_text(text[match.end() :]).strip(" -:.,;()[]")
        trailing = re.sub(r"^(college\s*:?)\s*", "", trailing, flags=re.IGNORECASE)

        if trailing:
            known_names[code] = trailing
            return code, trailing

        for near_idx in range(max(0, idx - 2), min(len(row_values), idx + 8)):
            if near_idx == idx:
                continue
            candidate = clean_text(row_values[near_idx])
            if not candidate:
                continue
            if _RE_ECODE.search(candidate):
                continue
            if candidate.upper() in CATEGORY_CODES or _RE_NUMERIC.match(candidate):
                continue
            known_names[code] = candidate
            return code, candidate

        return code, known_names.get(code, f"College {code}")
    return None


def extract_source_from_xlsx() -> Tuple[set, dict]:
    keys = set()
    stats = {}
    known_names: Dict[str, str] = {}

    for path in XLSX_FILES:
        year = year_from_filename(path.name)
        round_value = round_from_xlsx_filename(path.name)
        if not year:
            continue

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        current_code: Optional[str] = None
        current_category_map: Dict[int, str] = {}
        file_rows = 0

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                values = list(row)
                college = extract_college_from_row(values, known_names)
                if college:
                    current_code = college[0]
                    current_category_map = {}
                    continue

                category_map = extract_category_map(values)
                if current_code and category_map:
                    current_category_map = category_map
                    continue

                if not current_code or not current_category_map:
                    continue

                course = extract_course_name(values, current_category_map.keys())
                if not course:
                    continue

                for col_idx, category in current_category_map.items():
                    if col_idx >= len(values):
                        continue
                    rank = parse_rank(values[col_idx])
                    if rank is None:
                        continue
                    keys.add((year, round_value, current_code, course, category, rank))
                    file_rows += 1

        workbook.close()
        stats[path.name] = {"rows": file_rows}

    return keys, stats


def extract_headers_from_pdf_page(page: pdfplumber.page.Page) -> List[dict]:
    words = page.extract_words(keep_blank_chars=True) or []
    lines: Dict[int, List[str]] = {}
    for word in words:
        top = round(float(word.get("top", 0)))
        lines.setdefault(top, []).append(word.get("text", ""))

    headers = []
    for top in sorted(lines.keys()):
        line = clean_text(" ".join(lines[top]))
        match = re.search(r"College\s*:\s*[\(\[]?([A-Z]\d{3})[\)\]]?\s*(.*)", line, re.IGNORECASE)
        if not match:
            continue
        code = match.group(1).upper()
        name = clean_text(match.group(2))
        if "Course Name" in name:
            name = clean_text(name.split("Course Name")[0])
        headers.append({"top": top, "code": code, "name": name})
    return headers


def extract_source_from_pdf() -> Tuple[set, dict]:
    keys = set()
    stats = {}

    for path in PDF_2025_FILES:
        round_value = round_from_pdf_filename(path.name)
        file_rows = 0
        current_code: Optional[str] = None

        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                headers = extract_headers_from_pdf_page(page)
                if headers:
                    current_code = max(headers, key=lambda h: h["top"])["code"]

                tables = page.find_tables() or []
                for table in tables:
                    top = table.bbox[1] if table.bbox else 0
                    candidates = [h for h in headers if h["top"] < top]
                    if candidates:
                        current_code = max(candidates, key=lambda h: h["top"])["code"]

                    if not current_code:
                        continue

                    table_rows = table.extract() or []
                    if not table_rows:
                        continue

                    category_map: Dict[int, str] = {}
                    header_idx = -1
                    for idx, row in enumerate(table_rows[:12]):
                        row_values = list(row) if row else []
                        detected = extract_category_map(row_values)
                        if len(detected) >= 8:
                            category_map = detected
                            header_idx = idx
                            break

                    if header_idx < 0 or not category_map:
                        continue

                    for row in table_rows[header_idx + 1 :]:
                        row_values = list(row) if row else []
                        course = clean_text(row_values[0] if row_values else "")
                        if not course or looks_like_non_course(course):
                            continue

                        for col_idx, category in category_map.items():
                            if col_idx >= len(row_values):
                                continue
                            rank = parse_rank(row_values[col_idx])
                            if rank is None:
                                continue
                            keys.add(("2025", round_value, current_code, course, category, rank))
                            file_rows += 1

        stats[path.name] = {"rows": file_rows}

    return keys, stats


def load_master_keys() -> Tuple[set, list]:
    payload = json.loads(MASTER_JSON.read_text(encoding="utf-8"))
    rows = payload["cutoffs"] if isinstance(payload, dict) else payload
    keys = set()
    for row in rows:
        year = str(row.get("year", ""))
        if year not in {"2023", "2024", "2025"}:
            continue
        keys.add(
            (
                year,
                str(row.get("round", "")),
                str(row.get("institute_code", "")).upper(),
                clean_text(row.get("course", "")),
                str(row.get("category", "")).upper(),
                int(row.get("cutoff_rank", 0)),
            )
        )
    return keys, rows


def write_diff_csv(path: Path, rows: List[tuple]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["year", "round", "institute_code", "course", "category", "cutoff_rank"])
        writer.writerows(rows)


def main() -> None:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    master_keys, master_rows = load_master_keys()
    xlsx_keys, xlsx_stats = extract_source_from_xlsx()
    pdf_keys, pdf_stats = extract_source_from_pdf()
    source_keys = xlsx_keys | pdf_keys

    missing = sorted(source_keys - master_keys)
    extra = sorted(master_keys - source_keys)

    missing_by_code = Counter((y, r, c) for (y, r, c, _, _, _) in missing)
    extra_by_code = Counter((y, r, c) for (y, r, c, _, _, _) in extra)

    report = {
        "master_total_rows": len(master_rows),
        "master_unique_keys": len(master_keys),
        "source_unique_keys": len(source_keys),
        "xlsx_source_rows": len(xlsx_keys),
        "pdf_source_rows": len(pdf_keys),
        "missing_in_master_count": len(missing),
        "extra_in_master_count": len(extra),
        "xlsx_file_stats": xlsx_stats,
        "pdf_file_stats": pdf_stats,
        "missing_by_college_top50": [
            {"year": y, "round": r, "institute_code": c, "count": n}
            for (y, r, c), n in missing_by_code.most_common(50)
        ],
        "extra_by_college_top50": [
            {"year": y, "round": r, "institute_code": c, "count": n}
            for (y, r, c), n in extra_by_code.most_common(50)
        ],
        "missing_samples": [
            {
                "year": y,
                "round": r,
                "institute_code": c,
                "course": course,
                "category": cat,
                "cutoff_rank": rank,
            }
            for (y, r, c, course, cat, rank) in missing[:200]
        ],
        "extra_samples": [
            {
                "year": y,
                "round": r,
                "institute_code": c,
                "course": course,
                "category": cat,
                "cutoff_rank": rank,
            }
            for (y, r, c, course, cat, rank) in extra[:200]
        ],
    }

    (REPORTS_DIR / "kcet_audit_report.json").write_text(
        json.dumps(report, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    write_diff_csv(REPORTS_DIR / "kcet_audit_missing.csv", missing)
    write_diff_csv(REPORTS_DIR / "kcet_audit_extra.csv", extra)

    print("Audit complete")
    print(f"  master unique keys: {len(master_keys):,}")
    print(f"  source unique keys: {len(source_keys):,}")
    print(f"  missing in master: {len(missing):,}")
    print(f"  extra in master: {len(extra):,}")
    print(f"  report: {REPORTS_DIR / 'kcet_audit_report.json'}")


if __name__ == "__main__":
    main()
