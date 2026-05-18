#!/usr/bin/env python3
"""
KCET Comprehensive Data Extractor
==================================
Extracts ALL cutoff data from:
  - 2023 XLSX files (R1, R2, R3/Extended)
  - 2024 XLSX files (Mock, R1, R2, R3/Extended)
  - 2025 PDF files (Mock, R1, R2, R3) -- strictly from PDFs

Outputs both JSON and CSV. Target: 250,000+ entries.
"""

import os
import re
import csv
import json
import sys
from pathlib import Path
from collections import defaultdict

import openpyxl
import pdfplumber

# ─── Configuration ────────────────────────────────────────────────────────────

ROOT = Path(__file__).resolve().parent.parent
PUBLIC = ROOT / "public"
CUTOFFS_DIR = PUBLIC / "cutoffs"
OUTPUT_JSON = PUBLIC / "data" / "kcet_extracted_all.json"
OUTPUT_CSV = PUBLIC / "data" / "kcet_extracted_all.csv"

# All 24 KCET category codes
ALL_CATEGORIES = {
    '1G', '1K', '1R',
    '2AG', '2AK', '2AR',
    '2BG', '2BK', '2BR',
    '3AG', '3AK', '3AR',
    '3BG', '3BK', '3BR',
    'GM', 'GMK', 'GMR',
    'SCG', 'SCK', 'SCR',
    'STG', 'STK', 'STR',
}

# Load course code -> full name mapping
COURSE_MAPPING = {}
mapping_file = PUBLIC / "data" / "course_mapping.json"
if mapping_file.exists():
    with open(mapping_file, "r", encoding="utf-8") as f:
        COURSE_MAPPING = json.load(f)

# ─── Utility Functions ────────────────────────────────────────────────────────

def clean(text):
    """Normalize whitespace, trim, and remove control chars."""
    if text is None:
        return ""
    s = str(text)
    s = re.sub(r'[\r\n\t]+', ' ', s)
    s = re.sub(r'\s+', ' ', s)
    s = s.strip()
    return s


def parse_round_from_filename(filename):
    """Extract round from filename. R3 and EXT are unified as R3."""
    lower = filename.lower()
    if 'mock' in lower:
        return 'MOCK'
    if 'round3' in lower or 'extended' in lower or 'ext' in lower:
        return 'R3'
    if 'round2' in lower:
        return 'R2'
    if 'round1' in lower:
        return 'R1'
    return 'R1'


def parse_year_from_filename(filename):
    """Extract year from filename."""
    m = re.search(r'(202[0-9])', filename)
    return m.group(1) if m else None


def parse_cutoff(value):
    """Parse a cutoff rank value. Returns int or None."""
    if value is None:
        return None
    s = str(value).strip().replace(',', '')
    if s in ('', '--', '-', 'nan', 'None', 'NA', 'N/A'):
        return None
    try:
        n = int(float(s))
    except (ValueError, OverflowError):
        return None
    if n < 1 or n > 500000:
        return None
    return n


def make_key(row):
    """Composite dedup key."""
    return (
        row['institute_code'],
        row['course'],
        row['category'],
        row['year'],
        row['round'],
        str(row['cutoff_rank']),
    )


def resolve_course_name(raw_name, year):
    """
    For 2023/2024: if the raw name looks like a 2-letter code, expand via mapping.
    For 2025 PDFs: always keep the raw full name as-is.
    """
    if year == '2025':
        return clean(raw_name)

    cleaned = clean(raw_name)
    if not cleaned:
        return cleaned

    # Try to extract a course code prefix like "CS " or "CS-" or just "CS"
    # The XLSX format often has "CS Computer Science And Engineering" 
    # or just "CS" in the first column
    parts = re.split(r'\s+', cleaned, maxsplit=1)
    code = parts[0].upper()

    # If it's a known code, use the mapped full name
    if code in COURSE_MAPPING:
        return COURSE_MAPPING[code]

    # If the raw name is just a 2-3 letter code, try mapping
    if len(cleaned) <= 3 and cleaned.upper() in COURSE_MAPPING:
        return COURSE_MAPPING[cleaned.upper()]

    # Otherwise use the raw name (which may already be full)
    return cleaned


# ─── XLSX Extraction (2023 & 2024) ────────────────────────────────────────────

def find_xlsx_files(years):
    """Find all XLSX cutoff files for the given years."""
    files = []
    for f in sorted(PUBLIC.glob("kcet-*-cutoffs.xlsx")):
        fname = f.name.lower()
        for y in years:
            if str(y) in fname:
                files.append(f)
                break
    return files


def extract_college_info_from_sheet(ws):
    """
    Extract college code and name from the first few rows of an XLSX sheet.
    Patterns found:
      - "ENGINEERING - E001 College Name, Location"  (row 1 merged)
      - "E001" in one cell, name in adjacent cell
      - "E001 College Name" combined
    """
    for row_idx in range(1, min(ws.max_row + 1, 15)):
        for col_idx in range(1, min(ws.max_column + 1, 15)):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = clean(cell.value)
            if not val:
                continue

            # Pattern: "ENGINEERING - E001 College Name" or "E001 College Name"
            m = re.search(r'(E\d{3})\s+(.+)', val)
            if m:
                code = m.group(1)
                name = m.group(2).strip()
                # Remove trailing location after comma if present
                return code, name

            # Pattern: standalone "E001"
            m = re.match(r'^(E\d{3})$', val)
            if m:
                code = m.group(1)
                # Look for name in adjacent cells
                for next_col in range(col_idx + 1, min(col_idx + 10, ws.max_column + 1)):
                    next_val = clean(ws.cell(row=row_idx, column=next_col).value)
                    if next_val and len(next_val) > 3 and not re.match(r'^E\d{3}$', next_val):
                        return code, next_val
                # Check next row
                if row_idx < ws.max_row:
                    next_val = clean(ws.cell(row=row_idx + 1, column=col_idx).value)
                    if next_val and len(next_val) > 3:
                        return code, next_val
                return code, f"College {code}"

    return None, None


def find_category_header_row(ws):
    """
    Find the row that contains category column headers (GM, GMK, etc.).
    Returns (row_number, {category: column_index}) or (None, {}).
    """
    for row_idx in range(1, min(ws.max_row + 1, 30)):
        cat_map = {}
        for col_idx in range(1, ws.max_column + 1):
            val = clean(ws.cell(row=row_idx, column=col_idx).value).upper()
            if val in ALL_CATEGORIES:
                cat_map[val] = col_idx
        if len(cat_map) >= 8:  # Need at least 8 categories to be confident
            return row_idx, cat_map
    return None, {}


def find_course_column(ws, header_row):
    """
    Find which column contains course names (text content after the header row).
    """
    for col_idx in range(1, min(6, ws.max_column + 1)):
        course_count = 0
        for row_idx in range(header_row + 1, min(header_row + 20, ws.max_row + 1)):
            val = clean(ws.cell(row=row_idx, column=col_idx).value)
            if val and len(val) > 1 and re.search(r'[A-Za-z]', val) and val not in ('--', '-'):
                course_count += 1
        if course_count >= 2:
            return col_idx
    return 1  # Default to column A


def extract_xlsx_file(filepath, year, round_val):
    """Extract all cutoff data from a single XLSX file."""
    results = []
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Get college info
        code, name = extract_college_info_from_sheet(ws)
        if not code:
            continue

        # Find header row with categories
        header_row, cat_map = find_category_header_row(ws)
        if header_row is None:
            continue

        # Find course column
        course_col = find_course_column(ws, header_row)

        # Extract data rows
        for row_idx in range(header_row + 1, ws.max_row + 1):
            raw_course = clean(ws.cell(row=row_idx, column=course_col).value)
            if not raw_course:
                continue
            if raw_course in ('--', '-', 'nan'):
                continue
            low = raw_course.lower()
            if 'course' in low or 'branch' in low or 'programme' in low:
                continue
            if re.match(r'^[\d\s\-\.]+$', raw_course):
                continue

            course_name = resolve_course_name(raw_course, year)
            if not course_name:
                continue

            for cat, cat_col in cat_map.items():
                cell = ws.cell(row=row_idx, column=cat_col)
                rank = parse_cutoff(cell.value)
                if rank is None:
                    continue

                results.append({
                    'institute': name,
                    'institute_code': code,
                    'course': course_name,
                    'category': cat,
                    'cutoff_rank': rank,
                    'year': year,
                    'round': round_val,
                })

    wb.close()
    return results


def extract_all_xlsx(years=('2023', '2024')):
    """Extract from all XLSX files for the specified years."""
    files = find_xlsx_files(years)
    all_results = []

    for filepath in files:
        fname = filepath.name
        year = parse_year_from_filename(fname)
        round_val = parse_round_from_filename(fname)

        if year not in years:
            continue

        print(f"  📂 {fname} → Year={year}, Round={round_val}")
        rows = extract_xlsx_file(filepath, year, round_val)
        print(f"     ✅ {len(rows):,} entries extracted")
        all_results.extend(rows)

    return all_results


# ─── PDF Extraction (2025) ─────────────────────────────────────────────────────

def extract_pdf_file(pdf_path, year, round_val):
    """Extract all cutoff data from a single 2025 PDF file."""
    results = []

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        current_code = None
        current_name = None

        for page_idx, page in enumerate(pdf.pages):
            if page_idx % 50 == 0:
                print(f"     Page {page_idx}/{total_pages}...")

            # 1. Extract words and reconstruct lines to find college headers
            words = page.extract_words(keep_blank_chars=True)
            lines = {}
            for w in words:
                top = round(w['top'])
                if top not in lines:
                    lines[top] = []
                lines[top].append(w['text'])

            sorted_tops = sorted(lines.keys())
            college_headers = []

            for top in sorted_tops:
                line_text = " ".join(lines[top])
                # Match "College: (E001)Name" or "College : E001 Name"
                m = re.search(
                    r'College\s*:\s*[\(\[]?([A-Z0-9]{3,5})[\)\]]?\s*(.*)',
                    line_text,
                    re.IGNORECASE
                )
                if m:
                    c = m.group(1).strip()
                    n = m.group(2).strip()
                    # Clean up name: remove leading/trailing parens
                    n = re.sub(r'^[\(\)]+|[\(\)]+$', '', n).strip()
                    if len(c) >= 3:
                        college_headers.append({'top': top, 'code': c, 'name': n})

            # 2. Process tables on this page
            tables = page.find_tables()
            for table_obj in tables:
                table_top = table_obj.bbox[1]

                # Find closest college header above this table
                candidates = [h for h in college_headers if h['top'] < table_top]
                if candidates:
                    best = max(candidates, key=lambda x: x['top'])
                    current_code = best['code']
                    current_name = best['name']

                if not current_code:
                    continue

                table_data = table_obj.extract()
                if not table_data:
                    continue

                # Find category header row within this table
                cat_indices = {}
                for row in table_data:
                    if row is None:
                        continue
                    cleaned_row = [
                        str(c).strip().replace('\n', '') if c else ''
                        for c in row
                    ]
                    row_text = ' '.join(cleaned_row).upper()

                    # Check if this is a header row
                    if 'GM' in row_text and ('COURSE' in row_text or 'SCG' in row_text or '1G' in row_text):
                        for idx, cell in enumerate(cleaned_row):
                            cell_upper = cell.upper().strip()
                            if cell_upper in ALL_CATEGORIES:
                                cat_indices[cell_upper] = idx
                        continue

                    # If we haven't found category headers yet, skip
                    if not cat_indices:
                        # Maybe the first row IS the header without "Course" text
                        gm_count = sum(
                            1 for c in cleaned_row
                            if c.upper().strip() in ALL_CATEGORIES
                        )
                        if gm_count >= 8:
                            for idx, cell in enumerate(cleaned_row):
                                cell_upper = cell.upper().strip()
                                if cell_upper in ALL_CATEGORIES:
                                    cat_indices[cell_upper] = idx
                        continue

                    # Process data row
                    if len(row) == 0 or row[0] is None:
                        continue

                    course_raw = str(row[0]).strip().replace('\n', ' ')
                    if not course_raw or course_raw.upper() in ('COURSE', 'BRANCH', '--', '-', ''):
                        continue
                    if 'Course' in course_raw and 'GM' not in course_raw:
                        continue

                    course_name = resolve_course_name(course_raw, '2025')
                    if not course_name:
                        continue

                    for cat, idx in cat_indices.items():
                        if idx >= len(row):
                            continue
                        rank = parse_cutoff(row[idx])
                        if rank is None:
                            continue

                        results.append({
                            'institute': current_name,
                            'institute_code': current_code,
                            'course': course_name,
                            'category': cat,
                            'cutoff_rank': rank,
                            'year': year,
                            'round': round_val,
                        })

    return results


def extract_all_pdfs():
    """Extract from all 2025 PDF files."""
    pdf_files = sorted(CUTOFFS_DIR.glob("kcet-2025*.pdf"))
    all_results = []

    for pdf_path in pdf_files:
        fname = pdf_path.name
        round_val = parse_round_from_filename(fname)
        print(f"  📄 {fname} → Year=2025, Round={round_val}")
        rows = extract_pdf_file(pdf_path, '2025', round_val)
        print(f"     ✅ {len(rows):,} entries extracted")
        all_results.extend(rows)

    return all_results


# ─── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 70)
    print("  KCET COMPREHENSIVE DATA EXTRACTOR")
    print("  Target: 250,000+ entries | Output: JSON + CSV")
    print("=" * 70)
    print()

    all_data = []
    seen = set()

    # ── Phase 1: 2023 XLSX ──
    print("📥 Phase 1: Extracting 2023 XLSX data...")
    xlsx_2023 = extract_all_xlsx(years=('2023',))
    print(f"   → 2023 raw: {len(xlsx_2023):,}")
    print()

    # ── Phase 2: 2024 XLSX ──
    print("📥 Phase 2: Extracting 2024 XLSX data...")
    xlsx_2024 = extract_all_xlsx(years=('2024',))
    print(f"   → 2024 raw: {len(xlsx_2024):,}")
    print()

    # ── Phase 3: 2025 PDF ──
    print("📥 Phase 3: Extracting 2025 PDF data (strictly from PDFs)...")
    pdf_2025 = extract_all_pdfs()
    print(f"   → 2025 raw: {len(pdf_2025):,}")
    print()

    # ── Deduplication ──
    print("🔄 Deduplicating...")
    dupes = 0
    for row in xlsx_2023 + xlsx_2024 + pdf_2025:
        key = make_key(row)
        if key in seen:
            dupes += 1
            continue
        seen.add(key)
        all_data.append(row)

    print(f"   Duplicates removed: {dupes:,}")
    print(f"   Final unique entries: {len(all_data):,}")
    print()

    # ── Statistics ──
    by_year = defaultdict(int)
    by_round = defaultdict(int)
    institutes = set()
    courses = set()
    categories = set()

    for r in all_data:
        by_year[r['year']] += 1
        by_round[r['round']] += 1
        institutes.add(r['institute_code'])
        courses.add(r['course'])
        categories.add(r['category'])

    metadata = {
        'last_updated': __import__('datetime').datetime.now().isoformat(),
        'source_type': 'comprehensive_extraction',
        'total_entries': len(all_data),
        'total_institutes': len(institutes),
        'total_courses': len(courses),
        'total_categories': len(categories),
        'years_covered': sorted(by_year.keys(), reverse=True),
        'rounds_covered': sorted(by_round.keys()),
        'records_by_year': dict(sorted(by_year.items())),
        'records_by_round': dict(sorted(by_round.items())),
        'dedupe_key': 'institute_code+course+category+year+round+cutoff_rank',
    }

    # ── Save JSON ──
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    output = {
        'metadata': metadata,
        'cutoffs': all_data,
    }
    with open(OUTPUT_JSON, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False)
    print(f"💾 JSON saved: {OUTPUT_JSON}")
    json_size_mb = OUTPUT_JSON.stat().st_size / (1024 * 1024)
    print(f"   Size: {json_size_mb:.1f} MB")

    # ── Save CSV ──
    fieldnames = ['institute', 'institute_code', 'course', 'category', 'cutoff_rank', 'year', 'round']
    with open(OUTPUT_CSV, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_data)
    print(f"💾 CSV saved: {OUTPUT_CSV}")
    csv_size_mb = OUTPUT_CSV.stat().st_size / (1024 * 1024)
    print(f"   Size: {csv_size_mb:.1f} MB")
    print()

    # ── Summary ──
    print("=" * 70)
    print("  EXTRACTION SUMMARY")
    print("=" * 70)
    print(f"  📝 Total Entries:     {len(all_data):,}")
    print(f"  🏫 Unique Institutes: {len(institutes)}")
    print(f"  📚 Unique Courses:    {len(courses)}")
    print(f"  📊 Categories:        {len(categories)}")
    print()
    print("  📅 By Year:")
    for y, c in sorted(by_year.items()):
        print(f"     {y}: {c:,}")
    print()
    print("  🔄 By Round:")
    for r, c in sorted(by_round.items()):
        print(f"     {r}: {c:,}")
    print()

    if len(all_data) >= 250000:
        print("  ✅ SUCCESS: Exceeded 250,000 target!")
    else:
        print(f"  ⚠️  Below 250K target ({len(all_data):,}). Check source files.")

    print("=" * 70)


if __name__ == "__main__":
    main()
