#!/usr/bin/env python3
"""Debug XLSX structure - write results to JSON."""
import openpyxl
import re, json

def analyze_xlsx(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = {"file": path, "sheets": len(wb.sheetnames), "details": []}
    
    ALL_CATS = {'1G','1K','1R','2AG','2AK','2AR','2BG','2BK','2BR','3AG','3AK','3AR','3BG','3BK','3BR','GM','GMK','GMR','SCG','SCK','SCR','STG','STK','STR'}
    
    for sheet_name in wb.sheetnames[:3]:  # First 3 sheets
        ws = wb[sheet_name]
        sheet_info = {
            "name": sheet_name,
            "rows": ws.max_row,
            "cols": ws.max_column,
            "college_rows": [],
            "category_rows": [],
        }
        
        for row_idx in range(1, ws.max_row + 1):
            cat_count = 0
            row_cells = []
            for col_idx in range(1, min(ws.max_column + 1, 28)):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is not None:
                    s = str(v).replace('\n', ' ').strip()
                    row_cells.append({"col": col_idx, "val": s[:80]})
                    if s.upper() in ALL_CATS:
                        cat_count += 1
            
            row_text = " ".join(c["val"] for c in row_cells)
            
            m = re.search(r'(E\d{3,4}[A-Z]?)', row_text)
            if m and len(row_cells) <= 5:  # College header rows usually have few cells
                sheet_info["college_rows"].append({
                    "row": row_idx,
                    "code": m.group(1),
                    "text": row_text[:120]
                })
            
            if cat_count >= 10:
                sheet_info["category_rows"].append({
                    "row": row_idx,
                    "count": cat_count
                })
        
        result["details"].append(sheet_info)
    
    wb.close()
    return result

results = []
results.append(analyze_xlsx("public/kcet-2024-round1-cutoffs.xlsx"))
results.append(analyze_xlsx("public/kcet-2023-round1-cutoffs.xlsx"))

with open("scripts/xlsx_debug.json", "w") as f:
    json.dump(results, f, indent=2)
print("Done - saved to scripts/xlsx_debug.json")
